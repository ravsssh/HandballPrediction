import processingData
import settingData
import trainAnn
import predict
import openpyxl
import xlsxwriter
import os.path
import numpy as np


NUMBER_OF_SEASONS = 7
start_season_begin_year = 2014
start_season_end_year = 2015
isSettingDataEnabled = False

def getAnn():
    return ann

def winner(output):
    return max(enumerate(output), key=lambda x: x[1])[0]

def displayTestResult(results_ann, results_real):
    successfully_predicted = 0
    home_wins_total_predicted = 0
    draws_total_predicted  = 0
    away_wins_total_predicted  = 0
    home_wins_succ_predicted = 0
    draws_succ_predicted = 0
    away_wins_succ_predicted = 0
    for idx, res in enumerate(results_ann[0:,:]):
        if alphabet[winner(res)]==results_real[idx]:
            successfully_predicted = successfully_predicted + 1
        if (alphabet[winner(res)] == 1):
            home_wins_total_predicted = home_wins_total_predicted + 1
            if results_real[idx] == 1:
                home_wins_succ_predicted = home_wins_succ_predicted + 1
        if (alphabet[winner(res)] == 0):
            draws_total_predicted = draws_total_predicted + 1
            if results_real[idx] == 0:
                draws_succ_predicted = draws_succ_predicted + 1
        if (alphabet[winner(res)] == 2):
            away_wins_total_predicted = away_wins_total_predicted + 1
            if results_real[idx] == 2:
                away_wins_succ_predicted = away_wins_succ_predicted + 1
    
    print 'Successfully predicted games: ' + str(successfully_predicted*100/len(results_ann)) + '%'
    print 'Successfully predicted home wins: ' + str(home_wins_succ_predicted*100/home_wins_total_predicted) + '%'
    print 'home_wins_total_predicted: ', home_wins_total_predicted    
    if draws_total_predicted==0:
        print 'Successfully predicted draws: none'
    else:
        print 'Successfully predicted draws: ' + str(draws_succ_predicted*100/draws_total_predicted) + '%'
        print 'draws_total_predicted: ', draws_total_predicted   
    print 'Successfully predicted away wins: ' + str(away_wins_succ_predicted*100/away_wins_total_predicted) + '%'
    print 'away_wins_total_predicted: ', away_wins_total_predicted       
    return

def processData():
    
    if os.path.isfile('DataSet/Train/processedData.xlsx'):
        workbook = openpyxl.load_workbook("DataSet/Train/processedData.xlsx", use_iterators=True)
        sheet = workbook.worksheets[0]

        games = []
        for row in sheet.iter_rows():
            tempArray = []
            for idx,k in enumerate(row):
                tempArray.append(k.internal_value)
                
            games.append(tempArray)
    else:
        games, home_win, away_win, draw = processingData.startProcessingData(start_season_begin_year, start_season_end_year, NUMBER_OF_SEASONS, "Train")
        book = xlsxwriter.Workbook('DataSet/Train/processedData.xlsx')
        sheet = book.add_worksheet("Sheet 1")
        
        for i in range(0,len(games)):
            sheet.write(i,0,games[i][0])
            sheet.write(i,1,games[i][1])
            sheet.write(i,2,games[i][2])
            sheet.write(i,3,games[i][3])
            sheet.write(i,4,games[i][4])
            sheet.write(i,5,games[i][5])
            sheet.write(i,6,games[i][6])
            sheet.write(i,7,games[i][7])
        book.close()
    
    return games    
    
#writing data in Excel
if isSettingDataEnabled == True:
    for i in range(0,NUMBER_OF_SEASONS):
        seasonName = "DataSet/Train/season" + str(start_season_begin_year-i) + "-" + str(start_season_end_year-i) 
        list_of_teams,results = settingData.loadSeason(seasonName + ".html")
        settingData.writeSeasonInExcel(seasonName + ".xlsx",list_of_teams,results)

#traing ann
games, home_win, away_win, draw = processingData.startProcessingData(start_season_begin_year, start_season_end_year, NUMBER_OF_SEASONS, "Train")
#statistic_matrix = settingData.createStatisticMatrix(home_win, away_win, draw)
#settingData.printProcessedData(games, statistic_matrix)
#games = processData() #write games in excel file
print 'Processing train data is done!'
input_list, output_list = trainAnn.prepareDataForAnn(games)
output = trainAnn.convertOutput(output_list)
ann = trainAnn.create_ann(128,7,3)
ann = trainAnn.train_ann(ann, input_list, output)

#testing ann

'''for i in range(0,1):
        seasonName = "DataSet/Test/season" + str(2015-i) + "-" + str(2016-i) 
        list_of_teams,results = settingData.loadSeason(seasonName + ".html")
        settingData.writeSeasonInExcel(seasonName + ".xlsx",list_of_teams,results)'''
        
games, home_win, away_win, draw = processingData.startProcessingData(2015, 2016, 1, "Test")
input_list, output_list = trainAnn.prepareDataForAnn(games)
results_test = ann.predict(np.array(input_list, np.float32))

alphabet = [1, 0, 2]
displayTestResult(results_test, output_list)
predict.predictResult(2015,2016,21,ann)