import processingData
import openpyxl
import xlsxwriter
import numpy as np

NUMBER_OF_PREVIOUS_GAMES = processingData.NUMBER_OF_PREVIOUS_GAMES

def winner(output):
    return max(enumerate(output), key=lambda x: x[1])[0]

def predictResult(beginSeasonYear, endSeasonYear, roundNumber, ann):
    current_standing_coef, previous_standing1_coef, previous_standing2_coef, previous_standing3_coef, goal_difference, table, current_match_index = processingData.startProcessingData(2015, 2016, 1, "Predict")
    
    #processing_workbook = openpyxl.load_workbook("DataSet/Predict/season" + str(beginSeasonYear) + '-' + str(endSeasonYear) + "round" + str(roundNumber) + ".xlsx" , use_iterators=True)
    #processing_sheet = processing_workbook.worksheets[0]       
    
    path = "DataSet/Predict/season" + str(beginSeasonYear) + '-' + str(endSeasonYear) + "round" + str(roundNumber) + ".xlsx"
    workbook = openpyxl.load_workbook(path, use_iterators=True)
    sheet = workbook.worksheets[0]    
    
    input_list = []
    
    alphabet = ['1', 'X', '2']
    home_list = []
    away_list = []
      
    for idx,row in enumerate(sheet.iter_rows()):
        home_team_name = row[0].internal_value.strip().encode('utf-8')
        home_list.append(home_team_name)
        away_team_name = row[1].internal_value.strip().encode('utf-8')
        away_list.append(away_team_name)
        home_team_strength = sum(goal_difference[home_team_name]) + current_standing_coef[home_team_name] + previous_standing1_coef[home_team_name] + previous_standing2_coef[home_team_name] + previous_standing3_coef[home_team_name]
        away_team_strength = sum(goal_difference[away_team_name]) + current_standing_coef[away_team_name] + previous_standing1_coef[away_team_name] + previous_standing2_coef[away_team_name] + previous_standing3_coef[away_team_name]
        
        temp = [0] * 7
        
        temp[0] = int(round(home_team_strength - away_team_strength,2))
        temp[1] = round(home_team_strength)
        temp[2] = round(away_team_strength)
        
        if current_match_index[home_team_name]==0:
            index_of_previous_match_home = NUMBER_OF_PREVIOUS_GAMES-1
        else:
            index_of_previous_match_home = current_match_index[home_team_name]-1
        if current_match_index[away_team_name]==0:
            index_of_previous_match_away = NUMBER_OF_PREVIOUS_GAMES-1
        else:
            index_of_previous_match_away = current_match_index[away_team_name]-1
            
        if index_of_previous_match_home >= 1:
            temp[3] = sum(goal_difference[home_team_name][index_of_previous_match_home-1:index_of_previous_match_home+1])
        else:
            temp[3] = goal_difference[home_team_name][index_of_previous_match_home] + goal_difference[home_team_name][NUMBER_OF_PREVIOUS_GAMES-1]
                    
        if index_of_previous_match_away >= 1:
            temp[4] = sum(goal_difference[away_team_name][index_of_previous_match_away-1:index_of_previous_match_away+1])
        else:
            temp[4] = goal_difference[away_team_name][index_of_previous_match_away] + goal_difference[away_team_name][NUMBER_OF_PREVIOUS_GAMES-1]

        temp[5] = round(current_standing_coef[home_team_name], 2)
        temp[6] = round(current_standing_coef[away_team_name], 2)
     
        input_list.append(temp)
        
    
    #use ann----------------------------
    results_test = ann.predict(np.array(input_list, np.float32))
    
    book = xlsxwriter.Workbook(path)
    sh = book.add_worksheet("Sheet 1")
    
    for i in range(0,len(results_test)):
        #processing_workbook.get_sheet(0).write(i,2,alphabet[winner(results_test[i])])
        #processing_sheet.write(i,2,alphabet[winner(results_test[i])])
        sh.write(i,0,home_list[i])
        sh.write(i,1,away_list[i])
        sh.write(i,2,alphabet[winner(results_test[i])])
        
    book.close()
    
    
