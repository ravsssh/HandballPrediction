import openpyxl
import numpy as np

from sklearn.cluster import KMeans




# PREPARE DATA SET FILES FOR PROCESSING

#NUMBER_OF_SEASONS = 7
NUMBER_OF_PREVIOUS_GAMES = 3
NUMBER_OF_FIRST_GAMES = 36 # Represent number of games where standing will not be included in match rating
                           # e.g. If league has 18 teams, and we won't included rankings for first 4 rounds that is 36 games(9*4)
MR_MAX = 300 # Match rating maximum
'''CURR_MAX_COEF = 20.0 # Max coefficient for current season
                     # e.g. best ranked team has coefficient 20, last team 2
PREV1_MAX_COEF = 15.0 
PREV2_MAX_COEF = 10.0
PREV3_MAX_COEF = 5.0'''

CURR_MAX_COEF = 60 # Max coefficient for current season
                     # e.g. best ranked team has coefficient 20, last team 2
PREV1_MAX_COEF = 15
PREV2_MAX_COEF = 8
PREV3_MAX_COEF = 3

'''CURR_MAX_COEF = 0.1 # Max coefficient for current season
                     # e.g. best ranked team has coefficient 20, last team 2
PREV1_MAX_COEF = 0.1
PREV2_MAX_COEF = 0.1
PREV3_MAX_COEF = 0.1'''


def calculateCurrentStandingCoefficient(table, current_standing_coef, number_of_clusters, max_coefficient, number_of_game, number_of_all_games):
    
    #sorted_table = collections.OrderedDict(sorted(table.values(),reverse=True))
    list_of_points = []
    list_of_clubs = []
    for key,value in table.iteritems():
        if table[key]>-1:
            list_of_points.append(value)
            list_of_clubs.append(key)
            
    list_of_points = np.array(list_of_points).reshape(len(list_of_points), 1) # prepare for k-means 
    k_means = KMeans(n_clusters=number_of_clusters, max_iter=2000, tol=0.00001, n_init=10)
    k_means.fit(list_of_points)
    
    
    copy_of_cluster_centers = []
    
    for i in range(0,number_of_clusters):
        copy_of_cluster_centers.append(k_means.cluster_centers_[i][0])
        
    copy_of_cluster_centers.sort(reverse=True)
    
    list_of_sorted_cluster_index = []
    
    for i in range(0,number_of_clusters):
        for j in range(0,number_of_clusters):
            if copy_of_cluster_centers[i]==k_means.cluster_centers_[j]:
                list_of_sorted_cluster_index.append(j)
    #sala mala
    max_coefficient = max_coefficient*number_of_game/number_of_all_games
    #kraj sale
    for idx,lab in enumerate(k_means.labels_):
        for i,val in enumerate(list_of_sorted_cluster_index):
            if val==lab:
                current_standing_coef[list_of_clubs[idx]] = max_coefficient - i*max_coefficient*1.0/number_of_clusters
            
    
    return current_standing_coef



def calculatePreviousStandingCoefficient(startYear, previous_standing_coef, number_of_clusters, max_coefficient, folder):
    season = "DataSet/" + folder + "/standing" + str(startYear) + "-" + str(startYear+1) + ".xlsx" 
    workbookPreviousStanding1 = openpyxl.load_workbook(season, use_iterators=True)
    sheetPreviousStanding1 = workbookPreviousStanding1.worksheets[0]
    
    list_of_points = []
    list_of_clubs = []
    for row in sheetPreviousStanding1.iter_rows():
            for idx,k in enumerate(row):
                if idx == 0:
                    list_of_clubs.append(k.internal_value.encode("utf-8"))
                else:
                    list_of_points.append(int(k.internal_value))
    
    list_of_points = np.array(list_of_points).reshape(len(list_of_points), 1) # prepare for k-means 
    k_means = KMeans(n_clusters=number_of_clusters, max_iter=2000, tol=0.00001, n_init=10)
    k_means.fit(list_of_points)
    
    curr_coef = max_coefficient
    
    # calculate coefficient for teams from corresponding .xlsx file based on number of points of that team
    for idx,lab in enumerate(k_means.labels_):
        if idx==0:
            value_of_prev_label = k_means.labels_[idx]
        else:
            if k_means.labels_[idx]!=value_of_prev_label:
                curr_coef = curr_coef - max_coefficient/(number_of_clusters*1.0)
                value_of_prev_label = k_means.labels_[idx]
                
        previous_standing_coef[list_of_clubs[idx]] = curr_coef
            
    
    return previous_standing_coef
    
    
    
def startProcessingData(current_season_begin_year, current_season_end_year, NUMBER_OF_SEASONS, folder):
         
        current_standing_coef = {} # represent team coefficients for current season(which will be necessary for calculating match rating)
                              # current_standing_coef is map, key is team name, value is team coefficient for current season
                              # for example: teams ranked 1-5 on table have coefficient 1
                              #              teams ranked 6-10 on table have coefficient 0.7 etc.
        previous_standing1_coef = {} # represent team coefficients for one season before current season
        previous_standing2_coef = {} # represent team coefficients for two season before current season
        previous_standing3_coef = {} # represent team coefficients for three season before current season
        
        goal_difference = {} # represent map where key is team name, and value is goal difference for that team in past
                             # 6(NUMBER_OF_PREVIOUS_GAMES) games
                             # for example: Team1 as host have next results in past 6 games: 20:10, 33:32, 15:20, 16:15, 28:12
                             # goal_difference["Team1"] = [10,1,-5,1,16]
        
        table = {} # represent current season table, key is team name, value number of points for current season
        
        number_of_games = 0 # Counter which represent number of game in the processing season
        
        games = [] # Represent matrix for match statistic for every game.
                     # Every match is represented in signle row
                     # Column 0 - match rating
                     # Column 1 - final score of game(0-draw,1-home team win, 2-away team win)
                     # Column 2 - home team strength - based on goal_difference and standing on table in current and previous seasons
                    #  Column 3 - away team strength
                     # Column 4 - goal difference for home team in previous n=2 games(for better prediction)
                    #  Column 5 - goal difference for away team in previous n=2 games
                     # Column 6 - coefficient for home team based on standing in current season
                    #  Column 7 - coefficient for away team based on standing in current season
                    
        current_match_index = {} # Represent map where key is team name, and value is position for writting in goal_difference
                                 # Range for value is in (0,NUMBER_OF_PREVIOUS_GAMES)
        
        home_win = [0] * MR_MAX*2 # represent number of home wins for specific match rating
                                  # e.g. if home team won game,and match rating for that game is 16 -> home_win[116] += 1
                                  # e.g. if home team won game,and match rating for that game is -100 -> home_win[0] += 1
        away_win = [0] * MR_MAX*2
        draw = [0] * MR_MAX*2
        
        workbook = openpyxl.load_workbook("DataSet/list_of_all_teams.xlsx", use_iterators=True)
        sheet = workbook.worksheets[0]
        
        
        # ------------------------INITIALISE NECESSARY FIELDS--------------------------------
        for row in sheet.iter_rows():
            for k in row:
                current_standing_coef[k.internal_value.strip().encode("utf-8")] = 0 
                previous_standing1_coef[k.internal_value.strip().encode("utf-8")] = 0
                previous_standing2_coef[k.internal_value.strip().encode("utf-8")] = 0
                previous_standing3_coef[k.internal_value.strip().encode("utf-8")] = 0
                goal_difference[k.internal_value.strip().encode("utf-8")] = [0] * NUMBER_OF_PREVIOUS_GAMES
                table[k.internal_value.strip().encode("utf-8")] = -1 # cleaning table
                current_match_index[k.internal_value.strip().encode('utf-8')] = 0
        
        
        
        
        current_season_begin_year = current_season_begin_year + 1
        current_season_end_year = current_season_end_year + 1
    
        for i in range(0,NUMBER_OF_SEASONS):
            print 'Starting processing season: ' + str(current_season_begin_year-NUMBER_OF_SEASONS+i) + '-' + str(current_season_end_year-NUMBER_OF_SEASONS+i) + ' ...' 
            
            if i == 0: 
                # if we are processing 1-st season calculate standing coefficients for previous seasons 
                
                #current_standing = calculateCurrentStanding()
                previous_standing1_coef = calculatePreviousStandingCoefficient(current_season_begin_year-NUMBER_OF_SEASONS-1,previous_standing1_coef,5,PREV1_MAX_COEF, folder)
                previous_standing2_coef = calculatePreviousStandingCoefficient(current_season_begin_year-NUMBER_OF_SEASONS-2,previous_standing2_coef,5,PREV2_MAX_COEF, folder)
                previous_standing3_coef = calculatePreviousStandingCoefficient(current_season_begin_year-NUMBER_OF_SEASONS-3,previous_standing3_coef,5,PREV3_MAX_COEF, folder)
            else:
                # if we aren't processing 1-st season convert standing coefficients based on previous coefficients
                # e.g. re-calculated current_standing_coef becomes previous_standing1_coef
                for key,value in previous_standing2_coef.iteritems():
                    previous_standing3_coef[key] = value*PREV3_MAX_COEF/PREV2_MAX_COEF
                for key,value in previous_standing1_coef.iteritems():
                    previous_standing2_coef[key] = value*PREV2_MAX_COEF/PREV1_MAX_COEF
                for key,value in current_standing_coef.iteritems():
                    previous_standing1_coef[key] = value*PREV1_MAX_COEF/CURR_MAX_COEF
                    
            
            processing_workbook = openpyxl.load_workbook("DataSet/"+folder+"/season" + str(current_season_begin_year-NUMBER_OF_SEASONS+i) + '-' + str(current_season_end_year-NUMBER_OF_SEASONS+i) + ".xlsx" , use_iterators=True)
            processing_sheet = processing_workbook.worksheets[0]            
            
            # ---------------------------------------------PROCESSING NEW SEASON----------------------------------------
            for idx,row in enumerate(processing_sheet.iter_rows()):
                if idx == 0:
                    for key,value in goal_difference.iteritems(): 
                        goal_difference[key] = [0] * NUMBER_OF_PREVIOUS_GAMES # Annul goal difference in new season
                        table[key] = -1
                        current_match_index[k.internal_value.strip().encode('utf-8')] = 0
                
                if(idx < NUMBER_OF_FIRST_GAMES):
                    for key,value in current_standing_coef.iteritems(): 
                        current_standing_coef[key] = 0
                else:
                    current_standing_coef = calculateCurrentStandingCoefficient(table, current_standing_coef, 5, CURR_MAX_COEF, processing_sheet.max_row, processing_sheet.max_row)
                
                home_team_name = row[0].internal_value.strip().encode('utf-8')
                away_team_name = row[1].internal_value.strip().encode('utf-8')
                home_team_strength = sum(goal_difference[home_team_name]) + current_standing_coef[home_team_name] + previous_standing1_coef[home_team_name] + previous_standing2_coef[home_team_name] + previous_standing3_coef[home_team_name]
                away_team_strength = sum(goal_difference[away_team_name]) + current_standing_coef[away_team_name] + previous_standing1_coef[away_team_name] + previous_standing2_coef[away_team_name] + previous_standing3_coef[away_team_name]
                
                index_MR = int(round(home_team_strength-away_team_strength+MR_MAX))

                games.append([0] * 8)
                games[number_of_games][0] = int(round(home_team_strength - away_team_strength,2))
                games[number_of_games][2] = round(home_team_strength)
                games[number_of_games][3] = round(away_team_strength)

                if current_match_index[home_team_name]==0:
                    index_of_previous_match_home = NUMBER_OF_PREVIOUS_GAMES-1
                else:
                    index_of_previous_match_home = current_match_index[home_team_name]-1
                if current_match_index[away_team_name]==0:
                    index_of_previous_match_away = NUMBER_OF_PREVIOUS_GAMES-1
                else:
                    index_of_previous_match_away = current_match_index[away_team_name]-1

                if index_of_previous_match_home >= 1:
                    games[number_of_games][4] = sum(goal_difference[home_team_name][index_of_previous_match_home-1:index_of_previous_match_home+1])
                else:
                    games[number_of_games][4] = goal_difference[home_team_name][index_of_previous_match_home] + goal_difference[home_team_name][NUMBER_OF_PREVIOUS_GAMES-1]
                    
                if index_of_previous_match_away >= 1:
                    games[number_of_games][5] = sum(goal_difference[away_team_name][index_of_previous_match_away-1:index_of_previous_match_away+1])
                else:
                    games[number_of_games][5] = goal_difference[away_team_name][index_of_previous_match_away] + goal_difference[away_team_name][NUMBER_OF_PREVIOUS_GAMES-1]

                games[number_of_games][6] = round(current_standing_coef[home_team_name], 2)
                games[number_of_games][7] = round(current_standing_coef[away_team_name], 2)

                if table[home_team_name] == -1:
                    table[home_team_name] = 0
                if table[away_team_name] == -1:
                    table[away_team_name] = 0

                if int(row[2].internal_value) > int(row[3].internal_value):
                    home_win[index_MR] = home_win[index_MR] + 1
                    games[number_of_games][1] = 1
                    table[home_team_name] = table[home_team_name] + 2
                elif int(row[2].internal_value) < int(row[3].internal_value):
                    away_win[index_MR] = away_win[index_MR] + 1
                    games[number_of_games][1] = 2
                    table[away_team_name] = table[away_team_name] + 2
                else:
                    draw[index_MR] = draw[index_MR] + 1
                    games[number_of_games][1] = 0
                    table[home_team_name] = table[home_team_name] + 1
                    table[away_team_name] = table[away_team_name] + 1

                goal_difference[home_team_name][current_match_index[home_team_name]] = int(row[2].internal_value) - int(row[3].internal_value)
                goal_difference[away_team_name][current_match_index[away_team_name]] = int(row[3].internal_value) - int(row[2].internal_value)

                current_match_index[home_team_name] = ((current_match_index[home_team_name]+1) % NUMBER_OF_PREVIOUS_GAMES)
                current_match_index[away_team_name] = ((current_match_index[away_team_name]+1) % NUMBER_OF_PREVIOUS_GAMES)
                
                number_of_games = number_of_games + 1
        
        print ' '
        print 'Processing ' + folder + ' seasons is DONE!'
        

        if folder == 'Predict':
            return current_standing_coef, previous_standing1_coef, previous_standing2_coef, previous_standing3_coef, goal_difference, table, current_match_index
        else:
            return games, home_win, away_win, draw
        
        
        
def printProcessedData(games, statistic_matrix):
    
    print "Number of games: ", len(games)
    counter = 0
    for row in games:
        #print row
        if(row[0] > 20 or row[0] < -20):
            counter = counter + 1
            #print row
    
    print "Number of MR> 20 or MR<20: " , counter
    
    for row in statistic_matrix:
        print row
    return
    
    
def createStatisticMatrix(home_win, away_win, draw):
    statistic_matrix = []
    counter = 0
    
    for i in range(0,len(home_win)):
        sum_of_scores = home_win[i]+draw[i]+away_win[i]
        if sum_of_scores != 0:
            statistic_matrix.append([0]*7)
            statistic_matrix[counter][0] = i-MR_MAX
            statistic_matrix[counter][1] = home_win[i]
            statistic_matrix[counter][2] = draw[i]
            statistic_matrix[counter][3] = away_win[i]
            statistic_matrix[counter][4] = round(home_win[i]*100.0/sum_of_scores,2)
            statistic_matrix[counter][5] = round(draw[i]*100.0/sum_of_scores,2)
            statistic_matrix[counter][6] = round(away_win[i]*100.0/sum_of_scores,2)
            counter = counter + 1
        
    return statistic_matrix